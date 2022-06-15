import os
from os import listdir
from os.path import isfile, join
import re
import sys
import operator
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

folder = ''

sub_folders = [name for name in os.listdir() if os.path.isdir(os.path.join("", name))]

print(sub_folders)
df=pd.DataFrame(sub_folders)

DataBASE=Workbook()
DataBASE.save("DATA.xlsx")
DataBASE = load_workbook("DATA.xlsx")
sheet1 = DataBASE.create_sheet('DATA-X1',0)
DATAX1 = DataBASE['DATA-X1']

for x in dataframe_to_rows(df):
    DATAX1.append(x)
    
DataBASE.save("DATA.xlsx")
print("All Done")