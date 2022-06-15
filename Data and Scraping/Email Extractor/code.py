
import re
import sys
import operator
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

#input here
data=" full text"
divider=";"



email_list_raw=data.split(divider)
email_list=list()

for email in email_list_raw:
	if email.find('@')>=1:
		if email.find('<')>=0:
			email_list.append(email[email.find('<')+1:email.find('>'):1])
			continue
		email_list.append(email)

df = pd.DataFrame(list(zip(email_list)), 
               columns =["Email"])

print(df.head())

DataBASE=Workbook()
DataBASE.save("DATA.xlsx")
DataBASE = load_workbook("DATA.xlsx")
sheet1 = DataBASE.create_sheet('DATA-X1',0)
DATAX1 = DataBASE['DATA-X1']

for x in dataframe_to_rows(df):
    DATAX1.append(x)
    
DataBASE.save("DATA.xlsx")
 
print("All Done")
